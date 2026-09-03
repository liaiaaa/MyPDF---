# -*- coding: utf-8 -*-
"""
测试脚本：构造样例 PDF 并验证所有核心功能。
相比早期版本，本脚本做内容级校验（不再只查"文件体积>500"）：
  · 用 CJK 字体（china-s）写中文，中文内容真实可读、可核验
  · 样例含"无边框表格"，验证表格识别（P1 修复）
  · 构造大图 PDF 验证压缩确实降尺寸（P2 修复）
  · 转 Excel/Word 后回读单元格/文本内容
"""
import os
import sys
import traceback

BASE = os.path.dirname(os.path.abspath(__file__))
TEST = os.path.join(BASE, "_test")
os.makedirs(TEST, exist_ok=True)
sys.path.insert(0, BASE)

import fitz
from PIL import Image, ImageDraw
import pdf_toolbox as tb

ok = 0
fail = 0


def check(name, fn):
    global ok, fail
    try:
        r = fn()
        if r:
            ok += 1
            print(f"[OK]   {name}  ->  {r if isinstance(r, str) else '成功'}")
        else:
            fail += 1
            print(f"[FAIL] {name}  ->  断言返回假值（未通过）")
    except Exception as e:
        fail += 1
        print(f"[FAIL] {name}  ->  {e}")
        traceback.print_exc()


# ---------- 构造测试素材 ----------
def make_sample_pdf(path, n=3):
    """含中文 + 无边框表格 + 内嵌图片的样例 PDF（CJK 字体，中文真实可读）"""
    doc = fitz.open()
    for i in range(n):
        page = doc.new_page(width=595, height=842)  # A4
        page.insert_text((72, 72), f"安全测试 PDF 第 {i+1} 页 （中文内容）",
                         fontsize=16, fontname="china-s")
        # 无边框表格（对齐排版，无任何线条）
        rows = [["序号", "项目", "数值"], ["1", "测试A", "100"], ["2", "测试B", "200"]]
        y = 120
        for r in rows:
            x = 72
            for c in r:
                page.insert_text((x, y), c, fontsize=12, fontname="china-s")
                x += 120
            y += 20
        # 内嵌图片
        img = Image.new("RGB", (300, 150), (200, 60, 60))
        d = ImageDraw.Draw(img)
        d.text((20, 60), f"IMG-{i+1}", fill=(255, 255, 255))
        pimg = os.path.join(TEST, f"_img{i+1}.png")
        img.save(pimg)
        page.insert_image(fitz.Rect(72, 200, 372, 350), filename=pimg)
    doc.save(path)
    doc.close()
    print(f"   已生成样例 PDF：{os.path.basename(path)} ({os.path.getsize(path)} 字节)")


p1 = os.path.join(TEST, "样例A.pdf")
p2 = os.path.join(TEST, "样例B.pdf")
make_sample_pdf(p1)
make_sample_pdf(p2)

# ---------- 1. 信息 + 中文可读 ----------
check("PDF 信息", lambda: tb.pdf_info(p1))
check("PDF 中文内容真实可读", lambda: "中文" in fitz.open(p1)[0].get_text() and "可读")

# ---------- 2. 合并 ----------
merged = os.path.join(TEST, "_out_merged.pdf")
check("PDF 合并", lambda: tb.pdf_merge([p1, p2], merged))
check("合并后页数=6", lambda: fitz.open(merged).page_count == 6 and "页数正确")

# ---------- 3. 拆分 ----------
out_split = os.path.join(TEST, "split")
os.makedirs(out_split, exist_ok=True)
check("PDF 拆分(每页)", lambda: tb.pdf_split(p1, out_split))
check("拆分文件数=3", lambda: len(os.listdir(out_split)) == 3)

# ---------- 4. 压缩：验证大图确实降尺寸 + RGB/灰度/CMYK 不损坏（P2） ----------
def _make_compress_fixture():
    """构造含 2000px RGB / 灰度 / CMYK 三类大图的 PDF"""
    doc = fitz.open()
    kinds = []
    # RGB 2000x1200
    Image.new("RGB", (2000, 1200), (80, 120, 200)).save(os.path.join(TEST, "_rgb.png"))
    kinds.append(os.path.join(TEST, "_rgb.png"))
    # 灰度 2000x1200
    Image.new("L", (2000, 1200), 128).save(os.path.join(TEST, "_gray.png"))
    kinds.append(os.path.join(TEST, "_gray.png"))
    # CMYK 2000x1200
    Image.new("CMYK", (2000, 1200), (0, 128, 128, 0)).save(os.path.join(TEST, "_cmyk.jpg"))
    kinds.append(os.path.join(TEST, "_cmyk.jpg"))
    for k in kinds:
        pg = doc.new_page(width=2000, height=1200)
        pg.insert_image(fitz.Rect(0, 0, 2000, 1200), filename=k)
    big_pdf = os.path.join(TEST, "_big.pdf")
    doc.save(big_pdf)
    doc.close()
    return big_pdf, kinds


big_pdf, _ = _make_compress_fixture()
comp = os.path.join(TEST, "_out_compressed.pdf")
check("PDF 压缩(中档, 含3类大图)", lambda: tb.pdf_compress(big_pdf, comp, "中"))
cdoc = fitz.open(comp)
c_sizes = []
for pg in cdoc:
    for im in pg.get_images(full=True):
        p = fitz.Pixmap(cdoc, im[0])
        c_sizes.append(p.width)
check(f"压缩后所有图已降尺寸(≤1800, 实际{sorted(set(c_sizes))})",
      lambda: all(w <= 1800 for w in c_sizes) and "已压缩")


def _center_pixels(doc):
    """取每页渲染图中心像素（校验颜色未被 Filter/ColorSpace 错配破坏）"""
    out = []
    for pg in doc:
        px = pg.get_pixmap()
        cx, cy = px.width // 2, px.height // 2
        out.append(px.pixel(cx, cy))
    return out


def _close(a, b, tol=40):
    return all(abs(x - y) <= tol for x, y in zip(a, b))


pix_rgb, pix_gray, pix_cmyk = _center_pixels(cdoc)
check(f"压缩后 RGB 图颜色正确(实{pix_rgb})",
      lambda: _close(pix_rgb, (80, 120, 200)) and "颜色OK")
check(f"压缩后灰度图颜色正确(实{pix_gray})",
      lambda: _close(pix_gray, (128, 128, 128)) and "颜色OK")
check(f"压缩后 CMYK 图颜色正确(实{pix_cmyk})",
      lambda: _close(pix_cmyk, (255, 128, 128), 60) and "颜色OK")
check("压缩后 PDF 可正常渲染", lambda: cdoc[0].get_pixmap().width > 0 and "渲染OK")
print(f"       压缩体积：{os.path.getsize(big_pdf)} -> {os.path.getsize(comp)} 字节")
cdoc.close()
check("普通PDF压缩", lambda: tb.pdf_compress(p1, os.path.join(TEST, "_c2.pdf"), "中"))

# ---------- 5. 加密 ----------
enc = os.path.join(TEST, "_out_encrypted.pdf")
check("PDF 加密(AES-256)", lambda: tb.pdf_encrypt(p1, enc, "123456"))
check("加密后需密码", lambda: fitz.open(enc).is_encrypted and "已加密")

# ---------- 6. 解密 ----------
dec = os.path.join(TEST, "_out_decrypted.pdf")
check("PDF 解密", lambda: tb.pdf_decrypt(enc, dec, "123456"))
check("解密后无需密码", lambda: not fitz.open(dec).is_encrypted)

# ---------- 7. PDF 转图片 ----------
out_imgs = os.path.join(TEST, "toimg")
os.makedirs(out_imgs, exist_ok=True)
check("PDF 转图片", lambda: tb.pdf_to_images(p1, out_imgs, dpi=100))
imgs = [f for f in os.listdir(out_imgs) if f.endswith(".png")]
check("转图片数量=3", lambda: len(imgs) == 3)

# ---------- 8. 图片转 PDF ----------
img2pdf = os.path.join(TEST, "_out_img2pdf.pdf")
check("图片转 PDF", lambda: tb.images_to_pdf([os.path.join(out_imgs, f) for f in imgs], img2pdf))
check("图片合成页数=3", lambda: fitz.open(img2pdf).page_count == 3)

# ---------- 9. PDF 转 Word：回读文本 + 图片 ----------
docx = os.path.join(TEST, "_out.docx")
check("PDF 转 Word", lambda: tb.pdf_to_word(p1, docx))
from docx import Document as _Doc
_w = _Doc(docx)
_wtxt = "\n".join(p.text for p in _w.paragraphs)
check("Word 含表格文本(测试A)", lambda: "测试A" in _wtxt and "内容正确")
check("Word 含页面图片", lambda: len(_w.inline_shapes) >= 1 and "图片已保留")

# ---------- 10. PDF 转 Excel：回读单元格（验证无边框表格，P1） ----------
xlsx = os.path.join(TEST, "_out.xlsx")
check("PDF 转 Excel", lambda: tb.pdf_to_excel(p1, xlsx))
from openpyxl import load_workbook
_x = load_workbook(xlsx)
_xws = _x.active
_xrows = [[_xws.cell(r, c).value for c in range(1, _xws.max_column + 1)]
          for r in range(1, _xws.max_row + 1)]
check(f"Excel 无边框表格识别为多列(max_column={_xws.max_column})",
      lambda: _xws.max_column >= 2 and "非单列糊化")
check("Excel 含表头(序号/项目)",
      lambda: any(any(v in (row or []) for row in _xrows) for v in ["序号", "项目"]) and "内容正确")

# ---------- 11. PDF 转 PPT ----------
pptx = os.path.join(TEST, "_out.pptx")
check("PDF 转 PPT", lambda: tb.pdf_to_ppt(p1, pptx, dpi=80))
from pptx import Presentation as _Prs
check("PPT 页数=3", lambda: len(list(_Prs(pptx).slides)) == 3)

# ---------- 12. 提取图片 ----------
out_ext = os.path.join(TEST, "extract")
os.makedirs(out_ext, exist_ok=True)
check("提取图片", lambda: tb.pdf_extract_images(p1, out_ext))
ext_imgs = os.listdir(out_ext)
check("提取图片数量>=3", lambda: len(ext_imgs) >= 3)

print(f"\n===== 测试结果：通过 {ok} 项，失败 {fail} 项 =====")
